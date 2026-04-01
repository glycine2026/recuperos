import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Procesador de Facturas", page_icon="📄", layout="wide")

st.title("📄 Procesador de Facturas")
st.markdown("Subí el Excel exportado de Monday y obtené **una fila por factura** con los cultivos agrupados.")

VALID_ESPECIES = {"maiz", "maíz", "soja", "trigo", "girasol", "cebada", "sorgo", "avena", "centeno"}

def is_valid_especie(especie_str):
    """Check if the string looks like a crop name, not a date range or summary."""
    if not especie_str or especie_str in ("Especie", "nan"):
        return False
    if "to" in especie_str.lower() or "-" in especie_str and any(c.isdigit() for c in especie_str):
        # Likely a date range like "2026-03-14 to 2026-03-30"
        lower = especie_str.lower()
        if lower.startswith("20") or "to 20" in lower:
            return False
    return True


def parse_invoices(file):
    df = pd.read_excel(file, sheet_name=None, header=None)
    sheet = list(df.values())[0]
    invoices = []
    current_invoice = None
    in_subitems = False

    for i in range(3, len(sheet)):
        row = sheet.iloc[i]
        col0 = row.iloc[0]
        col0_str = str(col0).strip() if pd.notna(col0) else ""

        # "Subitems" header row
        if col0_str == "Subitems":
            in_subitems = True
            continue

        # Subitem data rows (col0 is NaN)
        if pd.isna(col0):
            if current_invoice is not None and in_subitems:
                especie = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
                campana = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
                campo   = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ""

                if is_valid_especie(especie):
                    if especie not in current_invoice["_especies"]:
                        current_invoice["_especies"].append(especie)
                if campana and campana not in ("Campaña", "nan", "Campo/Establecimiento"):
                    # Guard against date-range summary rows
                    if not (campana.startswith("20") and len(campana) > 8):
                        if campana not in current_invoice["_campanas"]:
                            current_invoice["_campanas"].append(campana)
                if campo and campo not in ("Campo/Establecimiento", "nan"):
                    if campo not in current_invoice["_campos"]:
                        current_invoice["_campos"].append(campo)
            continue

        # Invoice (parent) row
        fecha_carga = row.iloc[2]
        cuit        = row.iloc[3]

        # Skip summary/footer rows (CUIT is 0 or non-numeric)
        try:
            cuit_float = float(str(cuit).replace(",", "").strip())
            if cuit_float == 0:
                continue
        except (ValueError, TypeError):
            continue

        if pd.notna(fecha_carga) and pd.notna(cuit):
            if current_invoice is not None:
                current_invoice["Especie(s)"]                  = " / ".join(current_invoice.pop("_especies"))
                current_invoice["Campaña(s)"]                  = " / ".join(current_invoice.pop("_campanas"))
                current_invoice["Campo(s)/Establecimiento(s)"] = " / ".join(current_invoice.pop("_campos"))
                invoices.append(current_invoice)

            in_subitems = False
            current_invoice = {
                "Número Factura":         col0_str,
                "Subelementos":           str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else "",
                "Fecha Carga":            fecha_carga,
                "CUIT":                   str(int(float(cuit))),
                "Código Padrón":          str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else "",
                "Razón Social":           str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else "",
                "Servicio Prestado":      str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else "",
                "Factura PDF":            str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else "",
                "Otros Adjuntos":         str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else "",
                "Estado":                 str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else "",
                "Titular Carta de Porte": str(row.iloc[10]).strip() if pd.notna(row.iloc[10]) else "",
                "Referencia Albor":       str(row.iloc[11]).strip() if pd.notna(row.iloc[11]) else "",
                "Estado Admin":           str(row.iloc[12]).strip() if pd.notna(row.iloc[12]) else "",
                "Fecha Vencimiento":      str(row.iloc[13]).strip() if pd.notna(row.iloc[13]) else "",
                "Fecha Ingreso":          row.iloc[14],
                "_especies":  [],
                "_campanas":  [],
                "_campos":    [],
            }

    if current_invoice is not None:
        current_invoice["Especie(s)"]                  = " / ".join(current_invoice.pop("_especies"))
        current_invoice["Campaña(s)"]                  = " / ".join(current_invoice.pop("_campanas"))
        current_invoice["Campo(s)/Establecimiento(s)"] = " / ".join(current_invoice.pop("_campos"))
        invoices.append(current_invoice)

    return pd.DataFrame(invoices)


def create_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas procesadas"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", start_color="EBF3FB")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = cell_border
    ws.row_dimensions[1].height = 30

    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.border    = cell_border
            cell.alignment = left_align
            if fill:
                cell.fill = fill

    col_widths = {
        "Número Factura":              20,
        "Subelementos":                35,
        "Fecha Carga":                 20,
        "CUIT":                        18,
        "Código Padrón":               18,
        "Razón Social":                32,
        "Servicio Prestado":           24,
        "Factura PDF":                 16,
        "Otros Adjuntos":              16,
        "Estado":                      16,
        "Titular Carta de Porte":      22,
        "Referencia Albor":            20,
        "Estado Admin":                16,
        "Fecha Vencimiento":           20,
        "Fecha Ingreso":               20,
        "Especie(s)":                  24,
        "Campaña(s)":                  16,
        "Campo(s)/Establecimiento(s)": 38,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── UI ───────────────────────────────────────────────────────────────────────

uploaded_file = st.file_uploader("📂 Subí el archivo Excel de Monday", type=["xlsx"])

if uploaded_file:
    with st.spinner("Procesando..."):
        try:
            result_df = parse_invoices(uploaded_file)
            st.success(f"✅ Se procesaron **{len(result_df)} facturas** correctamente.")

            st.subheader("Vista previa")
            preview_cols = [
                "Número Factura", "Razón Social", "CUIT",
                "Referencia Albor", "Especie(s)", "Campaña(s)",
                "Campo(s)/Establecimiento(s)", "Servicio Prestado",
                "Estado", "Estado Admin", "Fecha Vencimiento",
            ]
            st.dataframe(
                result_df[[c for c in preview_cols if c in result_df.columns]],
                use_container_width=True,
                height=420,
            )

            excel_bytes = create_excel(result_df)
            st.download_button(
                label="⬇️ Descargar Excel procesado",
                data=excel_bytes,
                file_name="facturas_procesadas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")
            st.exception(e)
else:
    st.info("👆 Subí el Excel exportado de Monday para comenzar.")

st.markdown("---")
st.caption("Procesador de Facturas · Bioceres")
